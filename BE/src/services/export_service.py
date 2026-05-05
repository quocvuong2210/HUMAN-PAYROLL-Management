"""
Export Service - Xuất dữ liệu ra Excel và PDF
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from io import BytesIO
from datetime import datetime

class ExportService:
    
    @staticmethod
    def export_employees_to_excel(employees):
        """
        Xuất danh sách nhân viên ra Excel
        
        Args:
            employees: List of employee dicts
        
        Returns:
            BytesIO: Excel file in memory
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Danh sách nhân viên"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            "STT", "Mã NV", "Họ và tên", "Email", "Số điện thoại",
            "Phòng ban", "Chức vụ", "Giới tính", "Ngày sinh",
            "Ngày vào làm", "Trạng thái"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data
        for row_num, emp in enumerate(employees, 2):
            ws.cell(row=row_num, column=1, value=row_num - 1).border = thin_border
            ws.cell(row=row_num, column=2, value=emp.get('EmployeeID')).border = thin_border
            ws.cell(row=row_num, column=3, value=emp.get('FullName')).border = thin_border
            ws.cell(row=row_num, column=4, value=emp.get('Email')).border = thin_border
            ws.cell(row=row_num, column=5, value=emp.get('PhoneNumber')).border = thin_border
            ws.cell(row=row_num, column=6, value=emp.get('DepartmentName')).border = thin_border
            ws.cell(row=row_num, column=7, value=emp.get('PositionName')).border = thin_border
            ws.cell(row=row_num, column=8, value=emp.get('Gender')).border = thin_border
            
            # Format dates
            dob = emp.get('DateOfBirth')
            if dob:
                ws.cell(row=row_num, column=9, value=str(dob).split('T')[0] if 'T' in str(dob) else str(dob)).border = thin_border
            else:
                ws.cell(row=row_num, column=9, value='').border = thin_border
            
            hire_date = emp.get('HireDate')
            if hire_date:
                ws.cell(row=row_num, column=10, value=str(hire_date).split('T')[0] if 'T' in str(hire_date) else str(hire_date)).border = thin_border
            else:
                ws.cell(row=row_num, column=10, value='').border = thin_border
            
            ws.cell(row=row_num, column=11, value=emp.get('Status')).border = thin_border
        
        # Auto-adjust column widths
        column_widths = [5, 10, 25, 30, 15, 20, 20, 12, 15, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def export_employees_to_pdf(employees):
        """
        Xuất danh sách nhân viên ra PDF
        
        Args:
            employees: List of employee dicts
        
        Returns:
            BytesIO: PDF file in memory
        """
        output = BytesIO()
        
        # Create PDF with landscape orientation
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=18,
        )
        
        # Container for elements
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # Title
        title = Paragraph("DANH SÁCH NHÂN VIÊN", title_style)
        elements.append(title)
        
        # Date
        date_text = f"Ngày xuất: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        date_para = Paragraph(date_text, styles['Normal'])
        elements.append(date_para)
        elements.append(Spacer(1, 12))
        
        # Table data
        data = [
            ['STT', 'Mã NV', 'Họ và tên', 'Email', 'Phòng ban', 'Chức vụ', 'Trạng thái']
        ]
        
        for idx, emp in enumerate(employees, 1):
            data.append([
                str(idx),
                str(emp.get('EmployeeID', '')),
                emp.get('FullName', ''),
                emp.get('Email', ''),
                emp.get('DepartmentName', ''),
                emp.get('PositionName', ''),
                emp.get('Status', '')
            ])
        
        # Create table
        table = Table(data, colWidths=[0.5*inch, 0.8*inch, 2*inch, 2.2*inch, 1.5*inch, 1.5*inch, 1.2*inch])
        
        # Table style
        table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Body
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            
            # Alternating rows
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        output.seek(0)
        
        return output
    
    @staticmethod
    def export_report_to_excel(report_data, report_type):
        """
        Xuất báo cáo ra Excel
        
        Args:
            report_data: Dict containing report data
            report_type: Type of report (salary, attendance, etc.)
        
        Returns:
            BytesIO: Excel file in memory
        """
        wb = Workbook()
        ws = wb.active
        ws.title = f"Báo cáo {report_type}"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"BÁO CÁO {report_type.upper()}"
        title_cell.font = Font(bold=True, size=16, color="1F4E78")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Date
        ws.merge_cells('A2:F2')
        date_cell = ws['A2']
        date_cell.value = f"Ngày xuất: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        date_cell.alignment = Alignment(horizontal="center")
        
        # Headers (row 4)
        headers = report_data.get('headers', [])
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data
        data_rows = report_data.get('data', [])
        for row_num, row_data in enumerate(data_rows, 5):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            max_length = 0
            column_letter = ws.cell(row=4, column=col_idx).column_letter
            
            # Check all cells in this column
            for row_idx in range(4, len(data_rows) + 5):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def export_report_to_pdf(report_data, report_type):
        """
        Xuất báo cáo ra PDF
        
        Args:
            report_data: Dict containing report data
            report_type: Type of report
        
        Returns:
            BytesIO: PDF file in memory
        """
        output = BytesIO()
        
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=18,
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=30,
            alignment=1
        )
        
        title = Paragraph(f"BÁO CÁO {report_type.upper()}", title_style)
        elements.append(title)
        
        # Date
        date_text = f"Ngày xuất: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        date_para = Paragraph(date_text, styles['Normal'])
        elements.append(date_para)
        elements.append(Spacer(1, 12))
        
        # Table
        headers = report_data.get('headers', [])
        data_rows = report_data.get('data', [])
        
        table_data = [headers] + data_rows
        
        # Calculate column widths
        num_cols = len(headers)
        col_width = 9.5 * inch / num_cols
        
        table = Table(table_data, colWidths=[col_width] * num_cols)
        
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(table)
        
        doc.build(elements)
        output.seek(0)
        
        return output

    @staticmethod
    def export_salary_to_excel(salaries, month):
        """
        Xuất bảng lương ra Excel
        
        Args:
            salaries: List of salary dicts
            month: Month string (YYYY-MM-DD)
        
        Returns:
            BytesIO: Excel file in memory
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Bảng lương"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        title_cell.value = f"BẢNG LƯƠNG THÁNG {month[:7] if month else datetime.now().strftime('%Y-%m')}"
        title_cell.font = Font(bold=True, size=16, color="1F4E78")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers (row 3)
        headers = [
            "STT", "Mã NV", "Họ và tên", "Phòng ban", "Chức vụ",
            "Lương cơ bản", "Thưởng", "Khấu trừ", "Thực nhận", "Trạng thái"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data
        for row_num, salary in enumerate(salaries, 4):
            ws.cell(row=row_num, column=1, value=row_num - 3).border = thin_border
            ws.cell(row=row_num, column=2, value=salary.get('EmployeeID')).border = thin_border
            ws.cell(row=row_num, column=3, value=salary.get('FullName')).border = thin_border
            ws.cell(row=row_num, column=4, value=salary.get('DepartmentName')).border = thin_border
            ws.cell(row=row_num, column=5, value=salary.get('PositionName')).border = thin_border
            
            # Format currency
            ws.cell(row=row_num, column=6, value=float(salary.get('BaseSalary', 0))).border = thin_border
            ws.cell(row=row_num, column=6).number_format = '#,##0'
            
            ws.cell(row=row_num, column=7, value=float(salary.get('Bonus', 0))).border = thin_border
            ws.cell(row=row_num, column=7).number_format = '#,##0'
            
            ws.cell(row=row_num, column=8, value=float(salary.get('Deductions', 0))).border = thin_border
            ws.cell(row=row_num, column=8).number_format = '#,##0'
            
            ws.cell(row=row_num, column=9, value=float(salary.get('NetSalary', 0))).border = thin_border
            ws.cell(row=row_num, column=9).number_format = '#,##0'
            ws.cell(row=row_num, column=9).font = Font(bold=True, color="006100")
            
            ws.cell(row=row_num, column=10, value=salary.get('salaryStatus', 'pending')).border = thin_border
        
        # Auto-adjust column widths
        from openpyxl.utils import get_column_letter
        column_widths = [5, 10, 25, 20, 20, 15, 15, 15, 15, 15]
        for i, width in enumerate(column_widths, 1):
            column_letter = get_column_letter(i)
            ws.column_dimensions[column_letter].width = width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def export_salary_to_pdf(salaries, month):
        """
        Xuất bảng lương ra PDF
        
        Args:
            salaries: List of salary dicts
            month: Month string
        
        Returns:
            BytesIO: PDF file in memory
        """
        output = BytesIO()
        
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=18,
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=30,
            alignment=1
        )
        
        title = Paragraph(f"BẢNG LƯƠNG THÁNG {month[:7] if month else datetime.now().strftime('%Y-%m')}", title_style)
        elements.append(title)
        
        # Date
        date_text = f"Ngày xuất: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        date_para = Paragraph(date_text, styles['Normal'])
        elements.append(date_para)
        elements.append(Spacer(1, 12))
        
        # Table data
        data = [
            ['STT', 'Mã NV', 'Họ và tên', 'Phòng ban', 'Lương CB', 'Thưởng', 'Khấu trừ', 'Thực nhận']
        ]
        
        for idx, salary in enumerate(salaries, 1):
            data.append([
                str(idx),
                str(salary.get('EmployeeID', '')),
                salary.get('FullName', ''),
                salary.get('DepartmentName', ''),
                f"{float(salary.get('BaseSalary', 0)):,.0f}",
                f"{float(salary.get('Bonus', 0)):,.0f}",
                f"{float(salary.get('Deductions', 0)):,.0f}",
                f"{float(salary.get('NetSalary', 0)):,.0f}"
            ])
        
        # Create table
        table = Table(data, colWidths=[0.5*inch, 0.8*inch, 2*inch, 1.5*inch, 1.2*inch, 1*inch, 1*inch, 1.3*inch])
        
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(table)
        doc.build(elements)
        output.seek(0)
        
        return output
    
    @staticmethod
    def export_attendance_to_excel(attendances, month):
        """
        Xuất bảng chấm công ra Excel
        
        Args:
            attendances: List of attendance dicts
            month: Month string (YYYY-MM-DD)
        
        Returns:
            BytesIO: Excel file in memory
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Bảng chấm công"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f"BẢNG CHẤM CÔNG THÁNG {month[:7] if month else datetime.now().strftime('%Y-%m')}"
        title_cell.font = Font(bold=True, size=16, color="1F4E78")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers (row 3)
        headers = [
            "STT", "Mã NV", "Họ và tên", "Phòng ban", "Chức vụ",
            "Ngày công", "Ngày nghỉ", "Ngày phép"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data
        for row_num, att in enumerate(attendances, 4):
            ws.cell(row=row_num, column=1, value=row_num - 3).border = thin_border
            ws.cell(row=row_num, column=2, value=att.get('EmployeeID')).border = thin_border
            ws.cell(row=row_num, column=3, value=att.get('FullName')).border = thin_border
            ws.cell(row=row_num, column=4, value=att.get('DepartmentName')).border = thin_border
            ws.cell(row=row_num, column=5, value=att.get('PositionName')).border = thin_border
            
            # Work days
            work_cell = ws.cell(row=row_num, column=6, value=int(att.get('WorkDays', 0)))
            work_cell.border = thin_border
            work_cell.font = Font(bold=True, color="006100")
            work_cell.alignment = Alignment(horizontal="center")
            
            # Absent days
            absent_cell = ws.cell(row=row_num, column=7, value=int(att.get('AbsentDays', 0)))
            absent_cell.border = thin_border
            absent_cell.font = Font(color="C00000")
            absent_cell.alignment = Alignment(horizontal="center")
            
            # Leave days
            leave_cell = ws.cell(row=row_num, column=8, value=int(att.get('LeaveDays', 0)))
            leave_cell.border = thin_border
            leave_cell.font = Font(color="FF6600")
            leave_cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        from openpyxl.utils import get_column_letter
        column_widths = [5, 10, 25, 20, 20, 12, 12, 12]
        for i, width in enumerate(column_widths, 1):
            column_letter = get_column_letter(i)
            ws.column_dimensions[column_letter].width = width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def export_attendance_to_pdf(attendances, month):
        """
        Xuất bảng chấm công ra PDF
        
        Args:
            attendances: List of attendance dicts
            month: Month string
        
        Returns:
            BytesIO: PDF file in memory
        """
        output = BytesIO()
        
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=18,
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=30,
            alignment=1
        )
        
        title = Paragraph(f"BẢNG CHẤM CÔNG THÁNG {month[:7] if month else datetime.now().strftime('%Y-%m')}", title_style)
        elements.append(title)
        
        # Date
        date_text = f"Ngày xuất: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        date_para = Paragraph(date_text, styles['Normal'])
        elements.append(date_para)
        elements.append(Spacer(1, 12))
        
        # Table data
        data = [
            ['STT', 'Mã NV', 'Họ và tên', 'Phòng ban', 'Chức vụ', 'Công', 'Nghỉ', 'Phép']
        ]
        
        for idx, att in enumerate(attendances, 1):
            data.append([
                str(idx),
                str(att.get('EmployeeID', '')),
                att.get('FullName', ''),
                att.get('DepartmentName', ''),
                att.get('PositionName', ''),
                str(int(att.get('WorkDays', 0))),
                str(int(att.get('AbsentDays', 0))),
                str(int(att.get('LeaveDays', 0)))
            ])
        
        # Create table
        table = Table(data, colWidths=[0.5*inch, 0.8*inch, 2*inch, 1.5*inch, 1.5*inch, 0.8*inch, 0.8*inch, 0.8*inch])
        
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(table)
        doc.build(elements)
        output.seek(0)
        
        return output
